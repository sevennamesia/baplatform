from django.shortcuts import render, redirect
from baweb import models
from django.http import JsonResponse
from ..forms.courseforms import CourseCommentForm, CourseForm
from ..forms.userforms import UserChangePasswordForm
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from ..utils.encrypt import md5

def course_list(request):
    info_dict = request.session.get('info')
    user = models.User.objects.filter(id=info_dict['id']).first()
    if user.type == 2:
        is_teacher = 1
        teacher = models.TeacherInfo.objects.filter(user=user).first()
        course_list = models.Course.objects.filter(teacher=teacher).all() 
    else: 
        is_teacher = 0
        course_list = []
        student = models.StudentInfo.objects.filter(user=user).first()
        studentcourse = models.StudentCourse.objects.filter(student=student).all()
        for obj in studentcourse:
            course_list.append(obj.course)
    return render(request, "course_list.html", {'is_teacher':is_teacher, "course_list":course_list})

@csrf_exempt 
def des_update(request, id):
    des_richtext = request.POST.get('description_richtext')
    models.Course.objects.filter(id=id).update(description_richtext=des_richtext)
    return JsonResponse({"status": True})

@csrf_exempt
def course_add(request):
    '''添加课程'''
    form = CourseForm(data=request.POST)
    #print(request.FILES)
    if form.is_valid():
        info_dict = request.session.get('info')
        user = models.User.objects.filter(id=info_dict['id']).first()
        teacher = models.TeacherInfo.objects.filter(user=user).first()
        obj = form.save(commit=False)
        obj.teacher = teacher
        if "course_profile_pic" in request.FILES:
            obj.course_profile_pic = request.FILES["course_profile_pic"]
        obj.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False})

@csrf_exempt
def course_page(request, id):
    info = request.session.get("info", "")
    if info == '':
        username = "来访者"
        user_id = -1
        is_login = False
        is_teacher = False
    else:
        username = info['name']
        user_id = info['id']
        is_login = True
        user = models.User.objects.filter(id=user_id).first()
        if user.type == 2:
            is_teacher = True
        else:
            is_teacher = False

    changepwd_form = UserChangePasswordForm 
    course = models.Course.objects.filter(id=id).first()
    content = { 
        "username": username, 
        "id": user_id,
        "changepwd_form":changepwd_form,
        "course":course,
        "is_login": is_login,
        "is_teacher": is_teacher
    }
    return render(request, 'course_page.html', content)

def course_delete(request, id):
    '''删除课程（级联删除）'''
    '''无权限'''
    ''''''
    course = models.Course.objects.filter(id=id).first()
    info_dict = request.session.get('info')
    user = models.User.objects.filter(id=info_dict['id']).first()
    if user.type == 2:
        if course.teacher.user != user:
            return redirect('/')
    else:
        return redirect('/')
    models.Course.objects.filter(id=id).delete()
    return redirect('/course')

@csrf_exempt
def student_import(request, id):

    course = models.Course.objects.filter(id=id).first()
    info_dict = request.session.get('info')
    user = models.User.objects.filter(id=info_dict['id']).first()
    if user.type == 2:
        if course.teacher.user != user:
            return JsonResponse({"status": False})
    else:
        return JsonResponse({"status": False})
    file_obj = request.FILES.get('file')
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]

    for row in sheet.iter_rows(min_row=5):
        username = str(row[1].value)
        name = str(row[2].value)
        exists = models.User.objects.filter(username=username).exists()
        if not exists:
            if row[1].value != None:
                user = models.User.objects.create(username=username, password=md5(username), type=1)
                student = models.StudentInfo.objects.create(user=user, name=name)
                models.StudentCourse.objects.create(student=student, course=course)
        else:
            user = models.User.objects.filter(username=username).first()
            student = models.StudentInfo.objects.filter(user=user).first() 
            models.StudentCourse.objects.create(student=student, course=course)
    return JsonResponse({"status": True})

@csrf_exempt
def student_add(request, id):
    course = models.Course.objects.filter(id=id).first()
    info_dict = request.session.get('info')
    user = models.User.objects.filter(id=info_dict['id']).first()
    if user.type == 2:
        if course.teacher.user != user:
            return JsonResponse({"status": False})
    else:
        return JsonResponse({"status": False})
    username = request.POST.get('username')
    exists = models.User.objects.filter(username=username).exists()
    if exists:
        user = models.User.objects.filter(username=username).first()
        student = models.StudentInfo.objects.filter(user=user).first() 
        models.StudentCourse.objects.create(student=student, course=course)
    else:
        user = models.User.objects.create(username=username, password=md5(username), type=1)
        student = models.StudentInfo.objects.create(user=user, name="未知")
        models.StudentCourse.objects.create(student=student, course=course)
    return JsonResponse({"status":True})

@csrf_exempt
def student_delete(request, id, sid):
    course = models.Course.objects.filter(id=id).first()
    info_dict = request.session.get('info')
    user = models.User.objects.filter(id=info_dict['id']).first()
    if user.type == 2:
        if course.teacher.user != user:
            return redirect('/')
    else:
        return redirect('/')
    user = models.User.objects.filter(id=sid).first()
    student = models.StudentInfo.objects.filter(user=user).first() 
    models.StudentCourse.objects.filter(student=student, course=course).delete()
    return redirect("/course/{}/student/list".format(id))

def student_list(request, id):
    info = request.session.get("info", "")
    username = info['name']
    user_id = info['id']
    changepwd_form = UserChangePasswordForm 
    course = models.Course.objects.filter(id=id).first()
    user = models.User.objects.filter(id=user_id).first()
    if user.type == 2:
        is_teacher = 1
        if course.teacher.user != user:
            return redirect('/')
    elif user.type == 1:
        is_teacher = 0
        return redirect('/')
    studentcourses = models.StudentCourse.objects.filter(course=course).all()
    student_list = []
    for st in studentcourses:
        student_list.append(st.student)
    content = { 
        "username": username, 
        "id": user_id,
        "changepwd_form":changepwd_form,
        "course":course,
        "student_list":student_list,
        "is_teacher": is_teacher,
    }
    return render(request, 'student_list.html', content)

def marks_list(request, id):
    '''显示所有打分情况， 如果未提交，默认分数为0'''
    course = models.Course.objects.filter(id=id).first()
    assignment_list = models.Assignment.objects.filter(course=course).all()
    info_dict = request.session.get('info')
    user = models.User.objects.filter(id=info_dict['id']).first()
    student = models.StudentInfo.objects.filter(user=user).first()
    marks_list = []
    for assignment in assignment_list:
        marks = []
        marks.append(assignment.id)
        marks.append(assignment.name)
        flag = 0
        if assignment.is_group:
            groupmember = models.GroupMember.objects.filter(student=student).first()
            group = groupmember.group
            groupmembers = models.GroupMember.objects.filter(group=group).all()
            for member in groupmembers:
                exists = models.AssignmentSubmit.objects.filter(assignment=assignment, student=member.student).exists()
                if exists:
                    submit = models.AssignmentSubmit.objects.filter(assignment=assignment, student=member.student).first()
                    flag = 1
                    marks.append(submit.marks)
                    marks.append(submit.max_marks)
        else:
            exists = models.AssignmentSubmit.objects.filter(assignment=assignment, student=student).exists()
            if exists:
                submit = models.AssignmentSubmit.objects.filter(assignment=assignment, student=student).first()
                flag = 1
                marks.append(submit.marks)
                marks.append(submit.max_marks)
        if flag == 0:
            marks.append(0)
            marks.append(100)
        marks_list.append(marks)
    return render(request, 'marks_list.html', {"marks_list":marks_list})

def comment(request, id):
    course = models.Course.objects.filter(id=id).first()
    info = request.session.get("info", "")
    username = info['name']
    user_id = info['id']
    user = models.User.objects.filter(id=user_id).first()
    if user.type == 2:
        teacher = models.TeacherInfo.objects.filter(user=user).first()
        is_teacher = True
        if course.teacher != teacher:
            return redirect('/')
    else:
        student = models.StudentInfo.objects.filter(user=user).first()
        exists = models.StudentCourse.objects.filter(student=student,course=course)
        is_teacher = False
        if not exists:
            return redirect('/')
    changepwd_form = UserChangePasswordForm 
    commentform = CourseCommentForm
    mycomment_list = models.CourseComment.objects.filter(course=course, user_id=user_id).all()
    comment_list = models.CourseComment.objects.filter(course=course).all()
    content = { 
        "username": username, 
        "id": user_id,
        "changepwd_form":changepwd_form,
        "course":course,
        "commentform": commentform, 
        "mycomment_list":mycomment_list,
        "comment_list":comment_list, 
        "is_teacher": is_teacher
    }
    return render(request, 'comment.html', content)

def teacher_courses(request, id):
    user = models.User.objects.filter(id=id).first()
    teacher = models.TeacherInfo.objects.filter(user=user).first()
    course_list = models.Course.objects.filter(teacher=teacher).all() 
    content = {
        "course_list":course_list,
        "is_login": False,
    }
    return render(request, "course.html", content)

def course_show(request, id):
    course = models.Course.objects.filter(id=id).first()
    content = {
        "course":course,
        "is_login": False,
    }
    return render(request, 'course_page.html', content)