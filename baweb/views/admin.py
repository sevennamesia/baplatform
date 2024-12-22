from baweb import models
from django.http import JsonResponse,HttpResponse
from django.shortcuts import redirect
from django.views.decorators.csrf import csrf_exempt

from openpyxl import load_workbook
from ..utils.encrypt import md5

@csrf_exempt
def teacher_import(request):
    info = request.session.get("info", "")
    user = models.User.objects.filter(id=info["id"]).first()
    if user.type != 3:
        return JsonResponse({"status": False, "msg": "没有权限"})

    file_obj = request.FILES.get('file')
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]

    for row in sheet.iter_rows(min_row=5):
        username = str(row[1].value)
        exists = models.User.objects.filter(username=username).exists()
        if not exists:
            if row[1].value != None:
                user = models.User.objects.create(username=username, password=md5(username), type=2)
                models.TeacherInfo.objects.create(user=user)
    return JsonResponse({"status": True})

@csrf_exempt
def teacher_add(request):
    info = request.session.get("info", "")
    user = models.User.objects.filter(id=info["id"]).first()
    if user.type != 3:
        return JsonResponse({"status": False, "msg": "没有权限"})
    
    username = request.POST.get('username')
    exists = models.User.objects.filter(username=username).exists()
    if exists:
        return JsonResponse({"status":False,"msg":"账号已存在"})
    
    user = models.User.objects.create(username=username, password=md5(username), type=2)
    models.TeacherInfo.objects.create(user=user)
    return JsonResponse({"status": True})

@csrf_exempt
def teacher_disabled(request, id):
    info = request.session.get("info", "")
    user = models.User.objects.filter(id=info['id']).first()
    if user.type != 3:
        return HttpResponse("没有权限")
    teacher = models.User.objects.filter(id=id).first()
    models.TeacherDisabled.objects.create(teacher=teacher)
    return redirect("/admin/")

@csrf_exempt
def teacher_abled(request, id):
    info = request.session.get("info", "")
    user = models.User.objects.filter(id=info["id"]).first()
    if user.type != 3:
        return HttpResponse("没有权限")
    teacher = models.User.objects.filter(id=id).first()
    models.TeacherDisabled.objects.filter(teacher=teacher).delete()
    return redirect("/admin/")

@csrf_exempt
def is_disabled(request, id):
    info = request.session.get("info", "")
    user = models.User.objects.filter(id=info['id']).first()
    if user.type != 3:
        return JsonResponse({"status": False, "msg": "没有权限"})
    teacher = models.User.objects.filter(id=id).first()
    exists = models.TeacherDisabled.objects.filter(teacher=teacher).exists()
    if exists:
        return JsonResponse({"status": True, "is_disabled":True,"id":id})
    else:
        return JsonResponse({"status": True, "is_disabled":False, "id":id})

def update_torder(request, id, order):
    '''
        更新teacher order
    '''
    info = request.session.get("info", "")
    user = models.User.objects.filter(id=info['id']).first()
    if user.type != 3:
        return JsonResponse({"status": False, "msg": "没有权限"})
    teacher = models.User.objects.filter(id=id).first()
    models.TeacherInfo.objects.filter(user=teacher).update(order=order)
    return JsonResponse({"status": True})

def update_corder(request, id, order):
    '''
        更新course order
    '''
    info = request.session.get("info", "")
    user = models.User.objects.filter(id=info['id']).first()
    if user.type != 3:
        return JsonResponse({"status": False, "msg": "没有权限"})
    models.Course.objects.filter(id=id).update(order=order)
    return JsonResponse({"status": True})
